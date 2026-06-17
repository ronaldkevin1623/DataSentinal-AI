from openpyxl import Workbook


def generate_audit_report(results):

    wb = Workbook()

    # =====================
    # Summary Sheet
    # =====================

    ws1 = wb.active
    ws1.title = "Summary"

    ws1.append(["Metric", "Value"])

    ws1.append(["Total Rows", results["total_rows"]])
    ws1.append(["Quality Score", results["quality_score"]])
    ws1.append(["Duplicate Rows", results["duplicate_rows"]])
    ws1.append(["Missing Values", results["missing_values"]])
    ws1.append(["Invalid Emails", results["invalid_emails"]])
    ws1.append(["Invalid Phones", results["invalid_phones"]])

    # =====================
    # Validation Sheet
    # =====================

    ws2 = wb.create_sheet("Validation Results")

    ws2.append(["Metric", "Value"])

    ws2.append([
        "Future Dates",
        results["future_dates"]
    ])

    ws2.append([
        "Invalid Order Dates",
        results["invalid_order_dates"]
    ])

    ws2.append([
        "Invalid Quantity",
        results["invalid_quantity"]
    ])

    ws2.append([
        "Invalid Price",
        results["invalid_price"]
    ])

    ws2.append([
        "Invalid Payment Modes",
        results["invalid_payment_modes"]
    ])

    # =====================
    # Column Analysis
    # =====================

    ws3 = wb.create_sheet("Column Analysis")

    ws3.append([
        "Column",
        "Data Type",
        "Missing Values",
        "Unique Values"
    ])

    for column in results["column_summary"]:

        ws3.append([
            column["column"],
            column["dtype"],
            column["missing_values"],
            column["unique_values"]
        ])

    output_path = "outputs/audit_report.xlsx"

    wb.save(output_path)

    return output_path